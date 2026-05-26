"""
pseudo_excel_streamlit.py — Pseudonymiseur Excel local, multi-feuilles
    pip install streamlit pandas openpyxl
    streamlit run pseudo_excel_streamlit.py
"""
import copy
from collections import Counter
import hmac
import json
import math
import re
import secrets
import hashlib
import io

import pandas as pd
import streamlit as st

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


st.set_page_config(layout="wide", page_title="Pseudonymiseur Pro", page_icon="🛡️")


# ============================================================
#  CONSTANTES
# ============================================================

STRATEGIES = {
    "keep":               "Ne rien faire (garder tel quel, formules préservées)",
    "sequential_id":      "Remplacer par ID_1, ID_2…",
    "scale_numeric":      "Multiplier par un facteur (map valeur par valeur)",
    "regex_replace":      "Remplacer une partie via regex",
    "pii_patterns":       "Detecter/remplacer e-mails, telephones, IBAN, SSN, cartes...",
    "hash_reversible":    "Hash stable reversible (stocke la correspondance dans la map)",
    "hash_non_reversible": "Hash stable non reversible (ne stocke pas l'original)",
    "hash_deterministic": "Empreinte stable legacy (reversible)",
    "date_shift":         "Décaler toutes les dates (map valeur par valeur)",
    "recalc_formula":     "Recalculer via formule custom",
}

STRATEGY_PARAMS = {
    "keep": [],
    "sequential_id":      [("prefix", "Préfixe", "ID", str)],
    "scale_numeric":      [("factor", "Facteur (vide = aléatoire 0.5–2.0)", "", float)],
    "regex_replace":      [("pattern", "Motif (regex)", r"\d+", str),
                           ("prefix",  "Préfixe",       "N",     str)],
    "pii_patterns":       [],
    "hash_reversible":    [("length",  "Longueur (4-64)", 16,   int)],
    "hash_non_reversible": [("length",  "Longueur (4-64)", 16,   int)],
    "hash_deterministic": [("length",  "Longueur (4–64)", 10,   int)],
    "date_shift":         [("days",    "Décalage en jours (vide = aléatoire ±1000)", "", int)],
    "recalc_formula":     [("formula", "Formule (ex: `HT`*(1+`TVA-rate`))", "", str)],
}

DATE_FORMATS_TRY = [
    "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%Y-%m-%d", "%Y/%m/%d",
    "%m/%d/%Y", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S",
    "%d %B %Y", "%d %b %Y",
]


# ============================================================
#  HELPERS TYPES / NaN
# ============================================================

def is_missing(v) -> bool:
    if v is None:
        return True
    try:
        if pd.isna(v):
            return True
    except (TypeError, ValueError):
        pass
    if isinstance(v, str) and v.strip().lower() in ("", "nan", "nat", "none", "<na>", "null"):
        return True
    return False


def coerce_numeric_smart(series: pd.Series) -> tuple[pd.Series, int]:
    if pd.api.types.is_numeric_dtype(series):
        return series.astype(float), 0

    def clean(v):
        if is_missing(v):
            return None
        s = str(v).strip()
        s = re.sub(r"[€$£¥\s\u00a0]", "", s)
        if not s:
            return None
        if "," in s and "." in s:
            if s.rfind(",") > s.rfind("."):
                s = s.replace(".", "").replace(",", ".")
            else:
                s = s.replace(",", "")
        elif "," in s:
            s = s.replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return None

    cleaned = series.map(clean)
    n_failed = cleaned.isna().sum() - series.map(is_missing).sum()
    return pd.to_numeric(cleaned, errors="coerce"), int(max(0, n_failed))


def _coerce_param(value, ptype, default=None):
    if value is None: return default
    if isinstance(value, str):
        v = value.strip()
        if v == "": return default
        if ptype is float:
            try: return float(v.replace(",", "."))
            except ValueError: return default
        if ptype is int:
            try: return int(float(v.replace(",", ".")))
            except ValueError: return default
        return v
    try: return ptype(value)
    except (ValueError, TypeError): return default


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    new_cols = []
    seen: dict[str, int] = {}
    for i, c in enumerate(df.columns):
        if c is None or (isinstance(c, float) and pd.isna(c)):
            name = f"_col_{i}"
        elif isinstance(c, str):
            name = c
        else:
            name = str(c)
        if name in seen:
            seen[name] += 1
            name = f"{name}.{seen[name]}"
        else:
            seen[name] = 0
        new_cols.append(name)
    df = df.copy()
    df.columns = new_cols
    return df


def normalize_col_name(name: str) -> str:
    return re.sub(r"\s+", "", str(name)).lower()


# ============================================================
#  CLÉS CANONIQUES POUR MAPS VALUE-BASED
# ============================================================

def canon_number_key(v) -> str | None:
    """Clé canonique pour une valeur numérique. None si non parseable."""
    if is_missing(v): return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            f = float(v)
            if pd.isna(f): return None
            return repr(f)  # repr préserve la précision IEEE 754
        except (ValueError, TypeError):
            return None
    s = str(v).strip()
    if not s: return None
    s = re.sub(r"[€$£¥\s\u00a0]", "", s)
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return repr(float(s))
    except ValueError:
        return None


def canon_date_key(v, hint_format: str | None = None) -> str | None:
    """Clé canonique pour une date (format ISO YYYY-MM-DD HH:MM:SS)."""
    if is_missing(v): return None
    if isinstance(v, pd.Timestamp):
        if pd.isna(v): return None
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if hasattr(v, "strftime") and not isinstance(v, str):
        try: return v.strftime("%Y-%m-%d %H:%M:%S")
        except Exception: return None
    s = str(v).strip()
    if not s: return None
    if hint_format:
        try:
            ts = pd.to_datetime(s, format=hint_format, errors="coerce")
            if pd.notna(ts): return ts.strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            pass
    try:
        dayfirst = not bool(re.match(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}", s))
        ts = pd.to_datetime(s, errors="coerce", dayfirst=dayfirst)
        if pd.notna(ts): return ts.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        pass
    return None


def _file_to_bytes(file_or_bytes) -> bytes:
    if hasattr(file_or_bytes, "getvalue"):
        return file_or_bytes.getvalue()
    if hasattr(file_or_bytes, "read"):
        data = file_or_bytes.read()
        try:
            file_or_bytes.seek(0)
        except Exception:
            pass
        return data
    return file_or_bytes


def _row_non_missing(raw: pd.DataFrame, row_idx: int) -> list:
    if row_idx < 0 or row_idx >= len(raw):
        return []
    return [v for v in raw.iloc[row_idx].tolist() if not is_missing(v)]


def _looks_like_header_label(v) -> bool:
    if is_missing(v):
        return False
    s = str(v).strip()
    if not s or len(s) > 80:
        return False
    if canon_number_key(s) is not None:
        return False
    if re.fullmatch(r"[-+]?[\d\s.,/%]+", s):
        return False
    if re.search(r"\d{1,4}[/\-.]\d{1,2}[/\-.]\d{1,4}", s):
        return False
    return True


def _header_row_score(raw: pd.DataFrame, row_idx: int) -> float:
    vals = _row_non_missing(raw, row_idx)
    if not vals:
        return -1.0
    labels = [str(v).strip() for v in vals]
    label_like = sum(1 for v in vals if _looks_like_header_label(v))
    unique_ratio = len(set(labels)) / max(1, len(labels))
    next_vals = _row_non_missing(raw, row_idx + 1)
    next_bonus = min(len(next_vals), len(vals)) * 0.25
    return len(vals) + label_like * 2.5 + unique_ratio + next_bonus


KNOWN_HEADER_LABELS = {
    "sheet", "column", "recommendedstrategy", "params", "why",
    "narrative", "expectedafterdecode", "metric", "value", "howitshoulddecode",
    "customername", "clientid", "invoiceref", "shiftedinvoicedate",
    "scaledgrossexposure", "txnref", "postingdate", "settlementdate",
    "manageremail", "comment", "amount", "amounteur", "amountht", "amountttc",
    "email", "phone", "address", "notes", "reference", "invoicedate",
    "birthdate", "iban", "ssn", "apikey", "sessiontoken", "riskscore", "status",
    "reason",
}


def _normalized_header_label(v) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(v).strip().lower())


def _header_name_hits(raw: pd.DataFrame, row_idx: int) -> int:
    vals = _row_non_missing(raw, row_idx)
    return sum(1 for v in vals if _normalized_header_label(v) in KNOWN_HEADER_LABELS)


def _probable_header_row(raw: pd.DataFrame, row_idx: int) -> bool:
    vals = _row_non_missing(raw, row_idx)
    if not vals:
        return False
    label_like = sum(1 for v in vals if _looks_like_header_label(v))
    if len(vals) == 1:
        below_wider = any(
            len(_row_non_missing(raw, r)) > 1
            for r in range(row_idx + 1, min(len(raw), row_idx + 6))
        )
        return label_like == 1 and not below_wider
    return label_like >= max(1, math.ceil(len(vals) * 0.5))


def _detect_table_layout(raw: pd.DataFrame) -> dict:
    if raw.empty:
        return {
            "header_row": None,
            "data_start_row": 0,
            "excel_column_indices": [],
            "generated_header": True,
        }

    first_non_empty = next((i for i in range(len(raw)) if _row_non_missing(raw, i)), 0)
    search_until = min(len(raw), first_non_empty + 30)

    named_header_candidates = []
    for i in range(first_non_empty, search_until):
        vals = _row_non_missing(raw, i)
        if not vals:
            continue
        hits = _header_name_hits(raw, i)
        if hits >= 2 or (len(vals) <= 3 and hits >= 1 and i + 1 < len(raw)):
            named_header_candidates.append((hits, len(vals), -i, i))
    if named_header_candidates:
        _, _, _, header_row = max(named_header_candidates)
        return {
            "header_row": header_row,
            "data_start_row": header_row + 1,
            "excel_column_indices": [
                c for c in range(raw.shape[1])
                if any(not is_missing(v) for v in [raw.iat[header_row, c]] + raw.iloc[header_row + 1:, c].tolist())
            ],
            "generated_header": False,
        }

    candidates = [
        (_header_row_score(raw, i), i)
        for i in range(first_non_empty, search_until)
    ]
    best_score, header_row = max(candidates, default=(-1.0, first_non_empty))
    if best_score < 0 or not _probable_header_row(raw, header_row):
        header_row = None
        data_start = first_non_empty
        generated_header = True
    else:
        data_start = header_row + 1
        generated_header = False

    active_cols = []
    for c in range(raw.shape[1]):
        if header_row is None:
            col_values = raw.iloc[data_start:, c].tolist()
        else:
            col_values = [raw.iat[header_row, c]] + raw.iloc[data_start:, c].tolist()
        if any(not is_missing(v) for v in col_values):
            active_cols.append(c)

    return {
        "header_row": header_row,
        "data_start_row": data_start,
        "excel_column_indices": active_cols,
        "generated_header": generated_header,
    }


def _dataframe_from_raw_sheet(raw: pd.DataFrame, layout: dict) -> pd.DataFrame:
    active_cols = layout.get("excel_column_indices", [])
    if not active_cols:
        return pd.DataFrame()

    header_row = layout.get("header_row")
    data_start = int(layout.get("data_start_row", 0))

    if header_row is None:
        columns = [f"_col_{c}" for c in active_cols]
    else:
        columns = []
        for pos, c in enumerate(active_cols):
            raw_name = raw.iat[header_row, c] if c < raw.shape[1] else None
            columns.append(f"_col_{pos}" if is_missing(raw_name) else str(raw_name).strip())

    body = raw.iloc[data_start:, active_cols].copy()
    body.columns = columns
    body = body.reset_index(drop=True)
    return normalize_columns(body)


def _excel_scalar(v):
    if is_missing(v):
        return None
    if isinstance(v, pd.Timestamp):
        return None if pd.isna(v) else v.to_pydatetime()
    if hasattr(v, "item") and not isinstance(v, (str, bytes)):
        try:
            return v.item()
        except Exception:
            return v
    return v


# ============================================================
#  LECTURE EXCEL AVEC FORMULES
# ============================================================

def read_excel_with_formulas(file_or_bytes) -> tuple[dict, dict, dict]:
    data = _file_to_bytes(file_or_bytes)

    excel = pd.ExcelFile(io.BytesIO(data))
    sheets_values: dict[str, pd.DataFrame] = {}
    sheet_layouts: dict[str, dict] = {}
    for sheet_name in excel.sheet_names:
        raw = pd.read_excel(
            io.BytesIO(data),
            sheet_name=sheet_name,
            header=None,
            dtype=object,
        )
        layout = _detect_table_layout(raw)
        sheets_values[sheet_name] = _dataframe_from_raw_sheet(raw, layout)
        sheet_layouts[sheet_name] = layout

    sheets_formulas: dict[str, dict] = {}
    if not HAS_OPENPYXL:
        return sheets_values, sheets_formulas, sheet_layouts

    try:
        wb = load_workbook(io.BytesIO(data), data_only=False)
    except Exception:
        return sheets_values, sheets_formulas, sheet_layouts

    for sheet_name, df in sheets_values.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        layout = sheet_layouts.get(sheet_name, {})
        data_start = int(layout.get("data_start_row", 1))
        excel_cols = layout.get("excel_column_indices", list(range(len(df.columns))))
        formulas = {}
        col_names = list(df.columns)
        for row_idx in range(len(df)):
            excel_row = data_start + 1 + row_idx
            for col_idx, col_name in enumerate(col_names):
                if col_idx >= len(excel_cols): break
                cell = ws.cell(row=excel_row, column=int(excel_cols[col_idx]) + 1)
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas[(row_idx, col_name)] = cell.value
        if formulas:
            sheets_formulas[sheet_name] = formulas
    return sheets_values, sheets_formulas, sheet_layouts


def write_workbook_with_formulas(sheets: dict, sheets_formulas: dict,
                                  config: dict,
                                  original_bytes: bytes | None = None,
                                  layouts: dict | None = None) -> bytes:
    if original_bytes and layouts and HAS_OPENPYXL:
        try:
            wb = load_workbook(io.BytesIO(original_bytes), data_only=False)
            for sheet_name, df in sheets.items():
                if sheet_name not in wb.sheetnames:
                    ws = wb.create_sheet(str(sheet_name)[:31] or "Sheet1")
                    for c_idx, col_name in enumerate(df.columns, start=1):
                        ws.cell(row=1, column=c_idx).value = str(col_name)
                    data_start = 1
                    excel_cols = list(range(len(df.columns)))
                else:
                    ws = wb[sheet_name]
                    layout = layouts.get(sheet_name, {})
                    data_start = int(layout.get("data_start_row", 1))
                    excel_cols = layout.get("excel_column_indices", list(range(len(df.columns))))

                sheet_cfg = (config or {}).get(sheet_name, {})
                formulas = sheets_formulas.get(sheet_name, {})
                col_names = list(df.columns)
                for r_pos, (_, row) in enumerate(df.iterrows()):
                    excel_row = data_start + 1 + r_pos
                    for c_pos, col_name in enumerate(col_names):
                        if c_pos >= len(excel_cols):
                            continue
                        excel_col = int(excel_cols[c_pos]) + 1
                        formula = formulas.get((r_pos, col_name))
                        strat = sheet_cfg.get(col_name, {}).get("strategy", "keep")
                        if formula and strat == "keep":
                            ws.cell(row=excel_row, column=excel_col).value = formula
                        else:
                            ws.cell(row=excel_row, column=excel_col).value = _excel_scalar(row[col_name])

            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()
        except Exception:
            pass

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        used_names = set()
        name_mapping = {}
        for name, df in sheets.items():
            safe_name = (str(name)[:31] or "Sheet1")
            base = safe_name
            i = 1
            while safe_name in used_names:
                suffix = f"_{i}"
                safe_name = base[:31 - len(suffix)] + suffix
                i += 1
            used_names.add(safe_name)
            name_mapping[name] = safe_name
            df.to_excel(writer, sheet_name=safe_name, index=False)

        wb = writer.book
        for sheet_name, df in sheets.items():
            if sheet_name not in sheets_formulas:
                continue
            actual_name = name_mapping.get(sheet_name)
            if not actual_name or actual_name not in wb.sheetnames:
                continue
            ws = wb[actual_name]
            sheet_cfg = (config or {}).get(sheet_name, {})
            col_names = list(df.columns)
            col_to_idx = {c: i for i, c in enumerate(col_names)}

            for (row_idx, col_name), formula in sheets_formulas[sheet_name].items():
                strat = sheet_cfg.get(col_name, {}).get("strategy", "keep")
                if strat != "keep":
                    continue
                if col_name not in col_to_idx:
                    continue
                excel_row = row_idx + 2
                excel_col = col_to_idx[col_name] + 1
                ws.cell(row=excel_row, column=excel_col).value = formula

    return buf.getvalue()


# ============================================================
#  HELPERS DATES
# ============================================================

def _try_parse_one(s: str, fmt: str) -> bool:
    try:
        result = pd.to_datetime(s, format=fmt, errors="coerce")
        return pd.notna(result)
    except (ValueError, TypeError):
        return False


def detect_date_format(series: pd.Series) -> str | None:
    samples = []
    for v in series.dropna().head(20):
        if is_missing(v): continue
        if isinstance(v, pd.Timestamp): return None
        if hasattr(v, "strftime") and not isinstance(v, str): return None
        samples.append(str(v))
    if not samples: return None
    best, best_score = None, 0
    for fmt in DATE_FORMATS_TRY:
        score = sum(1 for s in samples if _try_parse_one(s, fmt))
        if score > best_score:
            best, best_score = fmt, score
    return best if best_score >= len(samples) * 0.5 else None


def parse_dates_robust(series: pd.Series, hint_format: str | None = None) -> pd.Series:
    if hint_format:
        try: return pd.to_datetime(series, format=hint_format, errors="coerce")
        except Exception: pass
    fmt = detect_date_format(series)
    dayfirst = bool(fmt and fmt.startswith("%d"))
    try:
        return pd.to_datetime(series, errors="coerce", dayfirst=dayfirst)
    except Exception:
        return pd.Series([pd.NaT] * len(series), index=series.index)


def looks_like_date(v) -> bool:
    """Heuristique : la valeur ressemble-t-elle à une date ?"""
    if is_missing(v): return False
    if isinstance(v, pd.Timestamp): return True
    if hasattr(v, "strftime") and not isinstance(v, str): return True
    if isinstance(v, str):
        s = v.strip()
        if not s or len(s) < 6 or len(s) > 30: return False
        # au moins 2 séparateurs courants OU motif ISO
        return bool(re.search(r"\d{1,4}[/\-.]\d{1,2}[/\-.]\d{1,4}", s))
    return False


# ============================================================
#  HELPERS FORMULES
# ============================================================

def _make_placeholder(placeholders: dict, col: str) -> str:
    for ph, c in placeholders.items():
        if c == col: return ph
    ph = f"__COL_{len(placeholders)}__"
    placeholders[ph] = col
    return ph


def eval_formula(formula: str, df: pd.DataFrame) -> pd.Series:
    placeholders: dict[str, str] = {}
    used_cols: list[str] = []
    expr = formula

    for c in df.columns:
        token = f"`{c}`"
        if token in expr:
            ph = _make_placeholder(placeholders, str(c))
            expr = expr.replace(token, ph)
            used_cols.append(c)

    cols_str = sorted([str(x) for x in df.columns], key=len, reverse=True)
    for c in cols_str:
        if any(c == cx for cx in placeholders.values()): continue
        pattern = rf"(?<![A-Za-z0-9_]){re.escape(c)}(?![A-Za-z0-9_])"
        new_expr, n = re.subn(
            pattern,
            lambda m, col=c: _make_placeholder(placeholders, col),
            expr,
        )
        if n > 0:
            expr = new_expr
            used_cols.append(c)

    for ph, col in placeholders.items():
        expr = expr.replace(ph, f"__df__[{col!r}]")

    safe_df = df.copy()
    for c in set(used_cols):
        if c in safe_df.columns and not pd.api.types.is_numeric_dtype(safe_df[c]):
            safe_df[c], _ = coerce_numeric_smart(safe_df[c])

    try:
        return eval(expr, {"__builtins__": {}}, {"__df__": safe_df})
    except KeyError as e:
        raise ValueError(f"Colonne inexistante référencée par la formule : {e}")
    except Exception as e:
        raise ValueError(f"Erreur dans la formule '{formula}' : {e}")


PII_PATTERNS = [
    ("EMAIL", re.compile(r"(?<![\w.+-])[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}(?![\w.-])", re.I)),
    ("IBAN", re.compile(r"\b[A-Z]{2}\d{2}[A-Z0-9 ]{11,30}\b", re.I)),
    ("SSN", re.compile(r"\b\d{3}-\d{2}-\d{4}\b")),
    ("URL", re.compile(r"\bhttps?://[^\s<>\"]+", re.I)),
    ("IP", re.compile(r"\b(?:\d{1,3}\.){3}\d{1,3}\b")),
    ("CARD", re.compile(r"(?<!\d)(?:\d[ -]?){13,19}(?!\d)")),
    ("PHONE", re.compile(r"(?<!\w)\+?\d[\d .()/\\-]{7,}\d(?!\w)")),
]


def _digits_only(s: str) -> str:
    return re.sub(r"\D", "", s)


def _luhn_valid(s: str) -> bool:
    digits = [int(d) for d in _digits_only(s)]
    if not 13 <= len(digits) <= 19:
        return False
    total = 0
    parity = len(digits) % 2
    for i, d in enumerate(digits):
        if i % 2 == parity:
            d *= 2
            if d > 9:
                d -= 9
        total += d
    return total % 10 == 0


def _valid_pii_match(label: str, value: str) -> bool:
    token = value.strip()
    if not token:
        return False
    if label == "CARD":
        return _luhn_valid(token)
    if label == "PHONE":
        digits = _digits_only(token)
        if len(digits) < 8 or looks_like_date(token):
            return False
        return token.startswith("+") or bool(re.search(r"[ .()/\\-]", token))
    if label == "IP":
        parts = token.split(".")
        return len(parts) == 4 and all(p.isdigit() and 0 <= int(p) <= 255 for p in parts)
    return True


# ============================================================
#  MOTEUR — 100% value-based pour le reverse
# ============================================================

class PseudoEngine:
    """
    Map :
      {
        "_shared_lookups":   { "<col_canon>": {src: pseudo}, ... },     # strings
        "_number_map":       { "<repr(orig)>": pseudo_num, ... },       # GLOBAL value-based
        "_date_map":         { "<iso_orig>": "<iso_pseudo>", ... },     # GLOBAL value-based
        "_scale_factors":    { "<col_canon>": 1.234 },                  # fallback only
        "_date_shifts":      { "<col_canon>": 42 },                     # fallback only
        "_date_formats":     { "<col_canon>": "%d/%m/%Y" },             # pour réécriture
        "sheets":            { "<sheet>": { "<col>": {"strategy":...} } }
      }
    """

    def __init__(self, existing: dict | None = None):
        self.map = existing if existing else {}
        self.map.setdefault("_shared_lookups", {})
        self.map.setdefault("_number_map", {})
        self.map.setdefault("_number_aliases", {})
        self.map.setdefault("_number_context", {})
        self.map.setdefault("_date_map", {})
        self.map.setdefault("_date_aliases", {})
        self.map.setdefault("_date_context", {})
        self.map.setdefault("_scale_factors", {})
        self.map.setdefault("_date_shifts", {})
        self.map.setdefault("_date_formats", {})
        self.map.setdefault("_hash_salt", secrets.token_hex(32))
        self.map.setdefault("_non_reversible_hash_salt", secrets.token_hex(32))
        self.map.setdefault("_used_pseudonyms", {})
        self.map.setdefault("sheets", {})
        self._migrate_legacy_map()
        self._warnings: list[str] = []
        self._used_pseudonyms: set[str] = set()
        self._register_existing_pseudonyms()
        # Caches d'inverses pour le reverse
        self._inv_numbers: dict | None = None
        self._inv_numbers_by_col: dict | None = None
        self._inv_dates: dict | None = None
        self._inv_dates_by_col: dict | None = None
        self._inv_strings_global: dict | None = None

    def _migrate_legacy_map(self):
        sl = self.map.get("_shared_lookups", {})
        moved = []
        for k, v in list(sl.items()):
            if k.startswith("__factor__") and isinstance(v, dict) and "value" in v:
                col = k[len("__factor__"):]
                self.map["_scale_factors"][normalize_col_name(col)] = float(v["value"])
                moved.append(k)
            elif k.startswith("__date_shift__") and isinstance(v, dict) and "days" in v:
                col = k[len("__date_shift__"):]
                self.map["_date_shifts"][normalize_col_name(col)] = int(v["days"])
                moved.append(k)
        for k in moved:
            del sl[k]
        for sheet_cfg in self.map.get("sheets", {}).values():
            if not isinstance(sheet_cfg, dict):
                continue
            for entry in sheet_cfg.values():
                if isinstance(entry, dict) and entry.get("strategy") == "hash_deterministic":
                    entry["strategy"] = "hash_reversible"

    def _lookup_for(self, col: str) -> dict:
        return self.map["_shared_lookups"].setdefault(normalize_col_name(col), {})

    def _register_existing_pseudonyms(self):
        used = self.map.get("_used_pseudonyms", {})
        if isinstance(used, dict):
            self._used_pseudonyms.update(str(k) for k, v in used.items() if v)
        elif isinstance(used, list):
            self._used_pseudonyms.update(str(v) for v in used)

        for col_lu in self.map.get("_shared_lookups", {}).values():
            if not isinstance(col_lu, dict):
                continue
            for pseudo in col_lu.values():
                if isinstance(pseudo, str):
                    self._used_pseudonyms.add(pseudo)

        self.map["_used_pseudonyms"] = {p: True for p in sorted(self._used_pseudonyms)}

    def _mark_pseudonym_used(self, pseudo: str):
        self._used_pseudonyms.add(pseudo)
        self.map["_used_pseudonyms"][pseudo] = True

    def _reserve_pseudonym(self, candidate: str) -> str:
        candidate = str(candidate).strip() or "PX"
        if candidate not in self._used_pseudonyms:
            self._mark_pseudonym_used(candidate)
            return candidate
        i = 2
        while f"{candidate}_{i}" in self._used_pseudonyms:
            i += 1
        candidate = f"{candidate}_{i}"
        self._mark_pseudonym_used(candidate)
        return candidate

    def _next_token(self, prefix: str, lookup: dict) -> str:
        base = re.sub(r"[^A-Za-z0-9]+", "_", str(prefix or "PX").upper()).strip("_") or "PX"
        i = len(lookup) + 1
        while True:
            candidate = f"{base}_{i}"
            if candidate not in self._used_pseudonyms:
                self._mark_pseudonym_used(candidate)
                return candidate
            i += 1

    def _lookup_or_create_token(self, lookup: dict, src: str, prefix: str) -> str:
        if src in lookup:
            return lookup[src]
        pseudo = self._next_token(prefix, lookup)
        lookup[src] = pseudo
        return pseudo

    def _hash_pseudonym(self, value: str, length: int) -> str:
        salt = str(self.map.get("_hash_salt") or "")
        digest = hmac.new(salt.encode("utf-8"), value.encode("utf-8"), hashlib.sha256).hexdigest()
        return self._reserve_pseudonym(f"H_{digest[:length]}")

    def _non_reversible_hash_pseudonym(self, value: str, length: int) -> str:
        salt = str(self.map.get("_non_reversible_hash_salt") or "")
        digest = hmac.new(salt.encode("utf-8"), value.encode("utf-8"), hashlib.sha256).hexdigest()
        return self._reserve_pseudonym(f"NH_{digest[:length]}")

    def _pseudonymize_pii_text(self, value, lookup: dict):
        if is_missing(value):
            return None
        text = str(value)
        for label, rx in PII_PATTERNS:
            def repl(match, label=label):
                token = match.group(0)
                if not _valid_pii_match(label, token):
                    return token
                return self._lookup_or_create_token(lookup, token, label)
            text = rx.sub(repl, text)
        return text

    @staticmethod
    def _put_inverse(target: dict, pseudo_key: str, orig_key: str):
        if pseudo_key in target and target[pseudo_key] != orig_key:
            target[pseudo_key] = None
        else:
            target[pseudo_key] = orig_key

    def _first_lookup_inverse(self, lookup_name: str, pseudo: str) -> str | None:
        lookup = self.map.get("_shared_lookups", {}).get(lookup_name, {})
        if not isinstance(lookup, dict):
            return None
        matches = [src for src, mapped in lookup.items() if mapped == pseudo]
        return matches[0] if len(matches) == 1 else None

    def _reverse_invoice_ref(self, value: str) -> str:
        def repl(match):
            year = self._first_lookup_inverse("reference", match.group("year"))
            seq = self._first_lookup_inverse("reference", match.group("seq"))
            if year and seq:
                return f"INV-{year}-{seq}"
            return match.group(0)
        return re.sub(r"\bINV-(?P<year>REF_\d+)-(?P<seq>REF_\d+)\b", repl, value)

    def _reverse_labeled_ref(self, value: str) -> str:
        def repl(match):
            year = self._first_lookup_inverse("label", match.group("year"))
            seq = self._first_lookup_inverse("label", match.group("seq"))
            if year and seq:
                return f"INV-{year}-{seq}"
            return match.group(0)
        return re.sub(r"\bINV-(?P<year>LBL_\d+)-(?P<seq>LBL_\d+)\b", repl, value)

    def _reverse_composite_string(self, value: str) -> tuple[str, bool]:
        new_s = self._reverse_invoice_ref(value)
        new_s = self._reverse_labeled_ref(new_s)

        replacements = {}
        for pseudo, orig in (self._inv_strings_global or {}).items():
            if orig is not None:
                replacements[pseudo] = orig

        # In free text, common PII/hash tokens are safe to resolve from their owning lookup
        # even when another strategy reused the same numeric suffix elsewhere.
        targeted_lookups = (
            "email__pii", "phone__pii", "manageremail__pii", "comment__pii",
            "notes__pii", "apikey",
        )
        for lookup_name in targeted_lookups:
            lookup = self.map.get("_shared_lookups", {}).get(lookup_name, {})
            if not isinstance(lookup, dict):
                continue
            for src, pseudo in lookup.items():
                if isinstance(pseudo, str):
                    replacements.setdefault(pseudo, src)

        if replacements:
            pseudos = sorted(replacements.keys(), key=len, reverse=True)
            rx = re.compile("|".join(re.escape(p) for p in pseudos))
            new_s = rx.sub(lambda m: replacements[m.group(0)], new_s)

        return new_s, new_s != value

    def _build_inverses(self):
        if self._inv_numbers is None:
            # Inverse des nombres : pseudo_repr → repr(orig)
            self._inv_numbers = {}
            self._inv_numbers_by_col = {}
            for canon, col_map in self.map.get("_number_context", {}).items():
                col_inv = {}
                for orig_repr, pseudo in (col_map or {}).items():
                    pseudo_repr = canon_number_key(pseudo)
                    if pseudo_repr is not None:
                        self._put_inverse(col_inv, pseudo_repr, orig_repr)
                        self._put_inverse(self._inv_numbers, pseudo_repr, orig_repr)
                self._inv_numbers_by_col[canon] = col_inv
            for pseudo_repr, orig_repr in self.map.get("_number_aliases", {}).items():
                self._put_inverse(self._inv_numbers, pseudo_repr, orig_repr)
            for orig_repr, pseudo in self.map["_number_map"].items():
                pseudo_repr = canon_number_key(pseudo)
                if pseudo_repr is not None:
                    self._put_inverse(self._inv_numbers, pseudo_repr, orig_repr)
        if self._inv_dates is None:
            # Inverse des dates : iso_pseudo → iso_orig
            self._inv_dates = {}
            self._inv_dates_by_col = {}
            for canon, col_map in self.map.get("_date_context", {}).items():
                col_inv = {}
                for orig_iso, pseudo_iso in (col_map or {}).items():
                    self._put_inverse(col_inv, pseudo_iso, orig_iso)
                    self._put_inverse(self._inv_dates, pseudo_iso, orig_iso)
                self._inv_dates_by_col[canon] = col_inv
            for pseudo_iso, orig_iso in self.map.get("_date_aliases", {}).items():
                self._put_inverse(self._inv_dates, pseudo_iso, orig_iso)
            for orig_iso, pseudo_iso in self.map["_date_map"].items():
                self._put_inverse(self._inv_dates, pseudo_iso, orig_iso)
        if self._inv_strings_global is None:
            # Inverse des strings : pseudo → orig (tous lookups confondus)
            self._inv_strings_global = {}
            for col_lu in self.map["_shared_lookups"].values():
                for src, pseudo in col_lu.items():
                    self._put_inverse(self._inv_strings_global, pseudo, src)

    # ---- pipeline apply ----

    def apply_sheet(self, df: pd.DataFrame, sheet: str, config: dict) -> pd.DataFrame:
        out = df.copy()
        sheet_map = self.map["sheets"].setdefault(sheet, {})

        for col, cfg in config.items():
            if col not in out.columns: continue
            strat = cfg.get("strategy", "keep")
            p = cfg.get("params", {}) or {}
            sheet_map[col] = {"strategy": strat, "params": p}
            canon = normalize_col_name(col)

            if strat in ("keep", "recalc_formula"):
                continue

            elif strat == "sequential_id":
                prefix = p.get("prefix") or "ID"
                lu = self._lookup_for(col)
                out[col] = pd.Series(
                    [(None if is_missing(v)
                      else self._lookup_or_create_token(lu, str(v).strip(), prefix))
                     for v in out[col]],
                    index=out.index, dtype="object",
                )

            elif strat == "scale_numeric":
                factor = _coerce_param(p.get("factor"), float)
                if factor is None or factor == 0:
                    factor = round(0.5 + secrets.randbelow(1500) / 1000, 4)
                self.map["_scale_factors"].setdefault(canon, factor)
                used_factor = self.map["_scale_factors"][canon]
                numeric, n_failed = coerce_numeric_smart(out[col])
                if n_failed > 0:
                    self._warnings.append(
                        f"Colonne '{col}' ({sheet}) : {n_failed} valeur(s) "
                        f"non numérique(s) conservée(s) telles quelles."
                    )
                res = []
                num_map = self.map["_number_map"]
                num_aliases = self.map["_number_aliases"]
                col_num_map = self.map["_number_context"].setdefault(canon, {})
                for orig_v, num_v in zip(out[col], numeric):
                    if is_missing(orig_v):
                        res.append(None); continue
                    if pd.isna(num_v):
                        res.append(orig_v); continue
                    orig_key = canon_number_key(orig_v)
                    if orig_key is None:
                        res.append(orig_v); continue
                    # Stocke la map valeur par valeur
                    if orig_key in col_num_map:
                        new_val = col_num_map[orig_key]
                    else:
                        new_val = float(num_v) * used_factor
                        col_num_map[orig_key] = new_val
                        num_map.setdefault(orig_key, new_val)
                    pseudo_key = canon_number_key(new_val)
                    if pseudo_key is not None:
                        num_aliases[pseudo_key] = orig_key
                    res.append(new_val)
                out[col] = pd.Series(res, index=out.index, dtype="object")

            elif strat == "regex_replace":
                pat = p.get("pattern") or r"\d+"
                pref = p.get("prefix") or "N"
                try:
                    rx = re.compile(pat)
                except re.error as e:
                    self._warnings.append(
                        f"Colonne '{col}' ({sheet}) : regex invalide ({e}), ignorée."
                    )
                    continue
                lu = self._lookup_for(col)
                out[col] = out[col].map(
                    lambda v: None if is_missing(v)
                    else rx.sub(
                        lambda m: self._lookup_or_create_token(lu, m.group(0), pref),
                        str(v),
                    )
                )

            elif strat == "pii_patterns":
                lu = self._lookup_for(f"{col}__pii")
                out[col] = out[col].map(lambda v: self._pseudonymize_pii_text(v, lu))

            elif strat in ("hash_reversible", "hash_deterministic"):
                length = _coerce_param(p.get("length"), int, default=10)
                if length is None or length < 4: length = 10
                if length > 64: length = 64
                lu = self._lookup_for(col)
                def hash_cell(v):
                    if is_missing(v):
                        return None
                    src = str(v).strip()
                    if src not in lu:
                        lu[src] = self._hash_pseudonym(src, length)
                    return lu[src]
                out[col] = out[col].map(hash_cell).astype("object")

            elif strat == "hash_non_reversible":
                length = _coerce_param(p.get("length"), int, default=16)
                if length is None or length < 4: length = 16
                if length > 64: length = 64
                anon_map = self.map.setdefault("_non_reversible_hashes", {}).setdefault(canon, {})
                def hash_cell(v):
                    if is_missing(v):
                        return None
                    src = str(v).strip()
                    fingerprint = hashlib.sha256(src.encode("utf-8")).hexdigest()
                    if fingerprint not in anon_map:
                        anon_map[fingerprint] = self._non_reversible_hash_pseudonym(src, length)
                    return anon_map[fingerprint]
                out[col] = out[col].map(hash_cell).astype("object")

            elif strat == "date_shift":
                days = _coerce_param(p.get("days"), int)
                if days is None:
                    days = secrets.randbelow(2000) - 1000
                self.map["_date_shifts"].setdefault(canon, days)
                used_days = self.map["_date_shifts"][canon]
                fmt = detect_date_format(out[col])
                self.map["_date_formats"][canon] = fmt
                parsed = parse_dates_robust(out[col], hint_format=fmt)

                non_missing = (~out[col].map(is_missing)).sum()
                n_parsed = parsed.notna().sum()
                if non_missing > 0 and n_parsed == 0:
                    self._warnings.append(
                        f"Colonne '{col}' ({sheet}) : aucune date reconnue, ignorée."
                    )
                    continue
                n_failed = int(non_missing - n_parsed)
                if n_failed > 0:
                    self._warnings.append(
                        f"Colonne '{col}' ({sheet}) : {n_failed} date(s) non reconnue(s)."
                    )

                shifted = parsed + pd.Timedelta(days=used_days)
                date_map = self.map["_date_map"]
                date_aliases = self.map["_date_aliases"]
                col_date_map = self.map["_date_context"].setdefault(canon, {})
                res = []
                for orig_v, sh in zip(out[col], shifted):
                    if is_missing(orig_v):
                        res.append(None); continue
                    if pd.isna(sh):
                        res.append(orig_v); continue
                    orig_key = canon_date_key(orig_v, hint_format=fmt)
                    pseudo_key = sh.strftime("%Y-%m-%d %H:%M:%S")
                    if orig_key:
                        col_date_map[orig_key] = pseudo_key
                        date_aliases[pseudo_key] = orig_key
                        date_map.setdefault(orig_key, pseudo_key)
                    res.append(sh.strftime(fmt) if fmt else sh)
                out[col] = pd.Series(res, index=out.index, dtype="object")

        # Recalculs
        for col, cfg in config.items():
            if col not in out.columns: continue
            if cfg.get("strategy") == "recalc_formula":
                formula = (cfg.get("params", {}) or {}).get("formula", "")
                if formula and formula.strip():
                    try:
                        out[col] = eval_formula(formula, out)
                    except ValueError as e:
                        self._warnings.append(f"Colonne '{col}' ({sheet}) : {e}")

        # Invalide les caches d'inverses (la map a changé)
        self._inv_numbers = self._inv_numbers_by_col = None
        self._inv_dates = self._inv_dates_by_col = None
        self._inv_strings_global = None
        return out

    def apply_workbook(self, sheets: dict, configs: dict) -> dict:
        return {n: self.apply_sheet(df, n, configs.get(n, {})) for n, df in sheets.items()}

    # ---- reverse 100% value-based ----

    @staticmethod
    def _num_value(v) -> float | None:
        key = canon_number_key(v)
        if key is None:
            return None
        try:
            return float(key)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _most_common_number(values: list[float], min_share: float = 0.6) -> float | None:
        clean = [v for v in values if v is not None and math.isfinite(v) and abs(v) > 1e-12]
        if not clean:
            return None
        rounded = [round(v, 10) for v in clean]
        best, count = Counter(rounded).most_common(1)[0]
        if count >= max(1, math.ceil(len(clean) * min_share)):
            return float(best)
        return None

    @staticmethod
    def _column_looks_scaled(col: str) -> bool:
        tokens = re.split(r"[\s_\-./]+", str(col).lower())
        blocked = {
            "id", "index", "rank", "count", "qty", "quantity", "rate", "ratio",
            "percent", "percentage", "pct", "tva", "vat", "age", "year", "month",
            "day", "date",
        }
        if any(t in blocked for t in tokens):
            return False
        markers = {
            "amount", "montant", "price", "prix", "total", "sum", "revenue",
            "sales", "cost", "salary", "balance", "value", "valeur", "budget",
            "ht", "ttc", "invoice", "facture", "payment", "paid",
        }
        return any(t in markers for t in tokens)

    @staticmethod
    def _column_looks_date(col: str) -> bool:
        tokens = re.split(r"[\s_\-./]+", str(col).lower())
        return any(t in {"date", "day", "month", "year", "period", "periode"} for t in tokens)

    def _infer_scale_factor(self, source: pd.Series, decoded: pd.Series, changed_flags: list[bool]) -> float | None:
        ratios = []
        for src, dec, changed in zip(source, decoded, changed_flags):
            if not changed:
                continue
            src_num = self._num_value(src)
            dec_num = self._num_value(dec)
            if src_num is None or dec_num in (None, 0):
                continue
            ratios.append(src_num / dec_num)
        return self._most_common_number(ratios)

    def _scale_factor_for_column(self, col: str, source: pd.Series,
                                 decoded: pd.Series, changed_flags: list[bool]) -> tuple[float | None, str]:
        inferred = self._infer_scale_factor(source, decoded, changed_flags)
        if inferred is not None:
            return inferred, "inferé depuis les valeurs déjà reconnues"

        canon = normalize_col_name(col)
        if canon in self.map["_scale_factors"]:
            return float(self.map["_scale_factors"][canon]), "nom de colonne connu"

        factors = [float(v) for v in self.map["_scale_factors"].values()
                   if v not in (None, 0)]
        uniq = sorted(set(round(v, 10) for v in factors if math.isfinite(v)))
        if len(uniq) == 1 and self._column_looks_scaled(col):
            return float(uniq[0]), "facteur unique du classeur"
        return None, ""

    def _apply_scale_reverse(self, col: str, source: pd.Series, decoded: pd.Series,
                             changed_flags: list[bool]) -> pd.Series:
        factor, reason = self._scale_factor_for_column(col, source, decoded, changed_flags)
        if factor in (None, 0) or abs(float(factor) - 1.0) < 1e-12:
            return decoded
        out = decoded.copy().astype(object)
        n = 0
        for idx, src, changed in zip(out.index, source, changed_flags):
            if changed:
                continue
            src_num = self._num_value(src)
            if src_num is None:
                continue
            out.at[idx] = src_num / float(factor)
            n += 1
        if n:
            self._warnings.append(
                f"Colonne '{col}' : {n} valeur(s) numérique(s) décodée(s) par facteur ({reason})."
            )
        return out

    def _infer_date_shift(self, source: pd.Series, decoded: pd.Series,
                          changed_flags: list[bool]) -> int | None:
        shifts = []
        for src, dec, changed in zip(source, decoded, changed_flags):
            if not changed:
                continue
            src_key = canon_date_key(src)
            dec_key = canon_date_key(dec)
            if not src_key or not dec_key:
                continue
            src_ts = pd.to_datetime(src_key, errors="coerce")
            dec_ts = pd.to_datetime(dec_key, errors="coerce")
            if pd.isna(src_ts) or pd.isna(dec_ts):
                continue
            delta = src_ts - dec_ts
            if delta.components.hours == delta.components.minutes == delta.components.seconds == 0:
                shifts.append(int(delta.days))
        if not shifts:
            return None
        best, count = Counter(shifts).most_common(1)[0]
        if count >= max(1, math.ceil(len(shifts) * 0.6)):
            return int(best)
        return None

    def _date_shift_for_column(self, col: str, source: pd.Series,
                               decoded: pd.Series, changed_flags: list[bool]) -> tuple[int | None, str]:
        inferred = self._infer_date_shift(source, decoded, changed_flags)
        if inferred is not None:
            return inferred, "inferé depuis les dates déjà reconnues"

        canon = normalize_col_name(col)
        if canon in self.map["_date_shifts"]:
            return int(self.map["_date_shifts"][canon]), "nom de colonne connu"

        shifts = [int(v) for v in self.map["_date_shifts"].values() if v is not None]
        uniq = sorted(set(shifts))
        if len(uniq) == 1 and self._column_looks_date(col):
            return int(uniq[0]), "décalage unique du classeur"
        return None, ""

    def _apply_date_reverse(self, col: str, source: pd.Series, decoded: pd.Series,
                            changed_flags: list[bool]) -> pd.Series:
        shift, reason = self._date_shift_for_column(col, source, decoded, changed_flags)
        if shift is None or shift == 0:
            return decoded
        out = decoded.copy().astype(object)
        n = 0
        for idx, src, changed in zip(out.index, source, changed_flags):
            if changed or not looks_like_date(src):
                continue
            src_key = canon_date_key(src)
            if not src_key:
                continue
            src_ts = pd.to_datetime(src_key, errors="coerce")
            if pd.isna(src_ts):
                continue
            out.at[idx] = src_ts - pd.Timedelta(days=int(shift))
            n += 1
        if n:
            self._warnings.append(
                f"Colonne '{col}' : {n} date(s) décodée(s) par décalage ({reason})."
            )
        return out

    def reverse_value(self, v, canon_col: str | None = None):
        """
        Tente de retrouver la valeur originale d'une cellule.
        Ordre : string global > number map > date map > composite string > inchangé.
        Retourne (nouvelle_valeur, a_changé).
        """
        if is_missing(v):
            return None, False

        if isinstance(v, str) and v in self._inv_strings_global:
            orig = self._inv_strings_global[v]
            if orig is not None:
                return orig, True

        num_key = canon_number_key(v)
        if num_key is not None:
            if canon_col and self._inv_numbers_by_col:
                orig_repr = self._inv_numbers_by_col.get(canon_col, {}).get(num_key)
                if orig_repr is not None:
                    try:
                        return float(orig_repr), True
                    except ValueError:
                        pass
            orig_repr = self._inv_numbers.get(num_key)
            if orig_repr is not None:
                try:
                    return float(orig_repr), True
                except ValueError:
                    pass

        if looks_like_date(v):
            date_key = canon_date_key(v)
            if date_key:
                if canon_col and self._inv_dates_by_col:
                    orig_iso = self._inv_dates_by_col.get(canon_col, {}).get(date_key)
                    if orig_iso is not None:
                        try:
                            return pd.to_datetime(orig_iso), True
                        except Exception:
                            return orig_iso, True
                orig_iso = self._inv_dates.get(date_key)
                if orig_iso is not None:
                    try:
                        return pd.to_datetime(orig_iso), True
                    except Exception:
                        return orig_iso, True

        if isinstance(v, str):
            new_s, changed = self._reverse_composite_string(v)
            if changed:
                return new_s, True

        return v, False

    def reverse_sheet(self, df: pd.DataFrame, sheet: str | None = None) -> pd.DataFrame:
        """Décodage value-based, avec inférence prudente pour feuilles transformées."""
        self._build_inverses()
        out = df.copy()

        for col in out.columns:
            canon = normalize_col_name(col)
            source_col = out[col].copy()
            new_col = []
            changed_flags = []
            for v in source_col:
                nv, changed = self.reverse_value(v, canon_col=canon)
                changed_flags.append(changed)
                new_col.append(nv)

            decoded = pd.Series(new_col, index=out.index, dtype="object")
            decoded = self._apply_date_reverse(col, source_col, decoded, changed_flags)
            decoded = self._apply_scale_reverse(col, source_col, decoded, changed_flags)

            fmt = self.map["_date_formats"].get(canon) or detect_date_format(source_col)
            if fmt:
                def reformat(v):
                    if isinstance(v, pd.Timestamp): return v.strftime(fmt)
                    return v
                decoded = decoded.map(reformat)

            out[col] = decoded

        return out

    def reverse_workbook(self, sheets: dict) -> dict:
        return {n: self.reverse_sheet(df, n) for n, df in sheets.items()}


# ============================================================
#  HELPERS UI
# ============================================================

def init_session():
    defaults = {
        "df_dict": None,
        "formulas_dict": None,
        "layout_dict": None,
        "workbook_bytes": None,
        "config": None,
        "map_data": None,
        "sheet_names": None,
        "file_name": None,
        "mode": "pseudonymiser",
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def suggest_strategy(col: str) -> str:
    cl = str(col).lower()
    tokens = re.split(r"[\s_\-./]+", cl)
    has = lambda *kws: any(t in kws for t in tokens)
    if "ttc" in tokens: return "recalc_formula"
    if has("tva", "rate", "taux", "vat"): return "keep"
    if has("ht", "montant", "prix", "amount", "price", "total"): return "scale_numeric"
    if has("email", "mail", "phone", "tel", "telephone", "mobile", "ssn", "iban",
           "address", "adresse", "comment", "comments", "note", "notes",
           "description", "text", "message"): return "pii_patterns"
    if has("id", "client", "nom", "name", "customer"): return "sequential_id"
    if has("date"): return "date_shift"
    if has("libelle", "libellé", "ref", "label", "reference"): return "regex_replace"
    return "keep"


def make_default_config(df_dict: dict, formulas_dict: dict) -> dict:
    cfg = {}
    for sheet, df in df_dict.items():
        cfg[sheet] = {}
        cols_with_formula = set()
        for (_, col_name) in formulas_dict.get(sheet, {}).keys():
            cols_with_formula.add(col_name)
        for c in df.columns:
            if c in cols_with_formula:
                cfg[sheet][c] = {"strategy": "keep", "params": {}}
            else:
                cfg[sheet][c] = {"strategy": suggest_strategy(c), "params": {}}
    return cfg


def smart_default(col: str, strat: str, all_cols: list, pname: str):
    cl = str(col).lower()
    tokens = re.split(r"[\s_\-./]+", cl)
    has = lambda *kws: any(t in kws for t in tokens)
    if strat == "sequential_id" and pname == "prefix":
        if has("client"): return "CLI"
        if has("nom", "name"): return "NOM"
        if has("id"): return "ID"
        return str(col).upper()[:5] if col else "ID"
    if strat == "recalc_formula" and pname == "formula" and "ttc" in tokens:
        ht = next((c for c in all_cols
                   if "ht" in re.split(r"[\s_\-./]+", str(c).lower())), "HT")
        tva = next((c for c in all_cols
                    if any(k in re.split(r"[\s_\-./]+", str(c).lower())
                           for k in ("tva", "rate", "taux", "vat"))), "TVA-rate")
        return f"`{ht}`*(1+`{tva}`)"
    if strat == "regex_replace":
        if pname == "pattern": return r"\d+"
        if pname == "prefix": return "N"
    return None


def merge_map_into_config(map_data: dict, config: dict) -> dict:
    if not map_data or not config: return config
    for sheet, sheet_cfg in map_data.get("sheets", {}).items():
        if sheet not in config: continue
        for col, entry in sheet_cfg.items():
            if col in config[sheet]:
                strat = entry.get("strategy", "keep")
                if strat in STRATEGIES:
                    config[sheet][col] = {
                        "strategy": strat,
                        "params": entry.get("params", {}) or {},
                    }
    return config


def safe_dataframe_for_display(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for c in out.columns:
        if out[c].dtype == "object":
            out[c] = out[c].map(lambda v: "" if is_missing(v) else str(v))
    return out


def has_formula_in_column(formulas_sheet: dict, col_name: str) -> int:
    return sum(1 for (_, c) in formulas_sheet.keys() if c == col_name)


# ============================================================
#  APP
# ============================================================

init_session()

st.title("🛡️ Pseudonymiseur Excel — local")
st.caption("Pseudonymisation **value-based** : la map stocke chaque correspondance, "
           "donc le décodage marche même sur des feuilles ajoutées par l'IA.")

mode = st.radio(
    "Mode",
    ["pseudonymiser", "decoder"],
    horizontal=True,
    format_func=lambda m: "🔒 Pseudonymiser" if m == "pseudonymiser" else "🔓 Décoder",
    key="mode",
)
st.divider()


# ============================================================
#  MODE DÉCODAGE
# ============================================================

if mode == "decoder":
    st.subheader("🔓 Décoder un fichier (potentiellement transformé par l'IA)")
    st.caption("Toutes les feuilles du fichier sont décodées, **y compris celles "
               "ajoutées par l'IA** (synthèses, pivots, etc.). Le décodage cherche "
               "chaque valeur dans la map, indépendamment du nom de feuille ou de colonne.")
    col1, col2 = st.columns(2)
    with col1:
        dec_excel = st.file_uploader("Fichier Excel", type=["xlsx"], key="dec_xl")
    with col2:
        dec_map = st.file_uploader("Map JSON associée", type=["json"], key="dec_map")

    if dec_excel and dec_map:
        try:
            dec_bytes = _file_to_bytes(dec_excel)
            sheets_in, formulas_in, layouts_in = read_excel_with_formulas(dec_bytes)
            map_data = json.load(dec_map)
            engine = PseudoEngine(map_data)
            result = engine.reverse_workbook(sheets_in)

            st.success(f"✓ Décodage effectué sur {len(result)} feuille(s).")
            if engine._warnings:
                with st.expander(f"⚠ {len(engine._warnings)} avertissement(s)"):
                    for w in engine._warnings:
                        st.warning(w)

            for name, df in result.items():
                with st.expander(f"📄 {name} ({len(df)} lignes)", expanded=False):
                    st.dataframe(safe_dataframe_for_display(df.head(50)),
                                 use_container_width=True, hide_index=True)

            decoded_bytes = write_workbook_with_formulas(
                result,
                formulas_in,
                {},
                original_bytes=dec_bytes,
                layouts=layouts_in,
            )
            st.download_button(
                "📥 Télécharger Excel décodé",
                data=decoded_bytes,
                file_name="decoded_output.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except Exception as e:
            st.error(f"Erreur : {e}")
    else:
        st.info("Fournis le fichier Excel **et** sa map JSON.")
    st.stop()


# ============================================================
#  MODE PSEUDONYMISATION
# ============================================================

with st.container():
    col_up1, col_up2 = st.columns(2)
    with col_up1:
        uploaded_excel = st.file_uploader("Fichier Excel à pseudonymiser (.xlsx)", type=["xlsx"])
    with col_up2:
        uploaded_map = st.file_uploader("Map existante à réutiliser (optionnel)", type=["json"])

if uploaded_excel and st.session_state.file_name != uploaded_excel.name:
    try:
        workbook_bytes = _file_to_bytes(uploaded_excel)
        df_dict, formulas_dict, layout_dict = read_excel_with_formulas(workbook_bytes)
    except Exception as e:
        st.error(f"Impossible de lire le fichier : {e}")
        st.stop()
    st.session_state.df_dict = df_dict
    st.session_state.formulas_dict = formulas_dict
    st.session_state.layout_dict = layout_dict
    st.session_state.workbook_bytes = workbook_bytes
    st.session_state.sheet_names = list(df_dict.keys())
    st.session_state.file_name = uploaded_excel.name
    st.session_state.config = make_default_config(df_dict, formulas_dict)
    for k in list(st.session_state.keys()):
        if (k.startswith("col_page_") or k.startswith("col_offset_")
            or k.startswith("filter_") or k.startswith("row_range_")
            or k.startswith("strat_") or k.startswith("param_")
            or k.startswith("offsetslider_") or k.startswith("prev_")
            or k.startswith("next_")):
            del st.session_state[k]
    for k in ("final_excel", "final_map", "final_warnings"):
        if k in st.session_state: del st.session_state[k]
    st.rerun()

if uploaded_map is not None:
    try:
        loaded = json.load(uploaded_map)
        st.session_state.map_data = loaded
        if st.session_state.config:
            st.session_state.config = merge_map_into_config(loaded, st.session_state.config)
    except Exception as e:
        st.error(f"Map JSON invalide : {e}")

if not st.session_state.df_dict:
    st.info("Charge un fichier Excel pour commencer.")
    st.stop()

st.divider()

sheet = st.sidebar.selectbox("📂 Feuille active", st.session_state.sheet_names)
df_current = st.session_state.df_dict[sheet]
formulas_current = (st.session_state.formulas_dict or {}).get(sheet, {})

n_formulas_total = len(formulas_current)
if n_formulas_total > 0:
    st.sidebar.info(f"📐 {n_formulas_total} formule(s) Excel détectée(s). "
                    "Préservées sur les colonnes 'keep'.")

col_gauche, col_droite = st.columns([1, 2], gap="large")


# ─── COLONNE GAUCHE ───
with col_gauche:
    st.subheader("🚀 Export")

    if st.button("🔒 Pseudonymiser tout le classeur",
                 use_container_width=True, type="primary"):
        with st.spinner("Traitement..."):
            try:
                final_engine = PseudoEngine(copy.deepcopy(st.session_state.map_data))
                processed = final_engine.apply_workbook(
                    st.session_state.df_dict,
                    st.session_state.config,
                )
                excel_bytes = write_workbook_with_formulas(
                    processed,
                    st.session_state.formulas_dict or {},
                    st.session_state.config,
                    original_bytes=st.session_state.workbook_bytes,
                    layouts=st.session_state.layout_dict,
                )
                map_bytes = json.dumps(
                    final_engine.map, indent=2, ensure_ascii=False, default=str
                ).encode("utf-8")
                st.session_state.final_excel = excel_bytes
                st.session_state.final_map = map_bytes
                st.session_state.final_warnings = list(final_engine._warnings)
                st.success("✓ Fichier généré.")
            except Exception as e:
                st.error(f"Erreur : {e}")

    if "final_excel" in st.session_state:
        d1, d2 = st.columns(2)
        with d1:
            st.download_button(
                "📥 Excel pseudo",
                data=st.session_state.final_excel,
                file_name="pseudonymise_output.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with d2:
            st.download_button(
                "📥 Map JSON",
                data=st.session_state.final_map,
                file_name="map_pseudo.json",
                mime="application/json",
                use_container_width=True,
            )
        if st.session_state.get("final_warnings"):
            with st.expander(f"⚠ {len(st.session_state.final_warnings)} avertissement(s)"):
                for w in st.session_state.final_warnings:
                    st.warning(w)

    st.divider()
    st.subheader("⚙️ Stratégies")
    st.caption("Modifications appliquées en direct.")

    scroll_config = st.container(height=650)
    with scroll_config:
        all_cols = list(df_current.columns)
        for c_name in all_cols:
            if not isinstance(c_name, str): continue
            if c_name not in df_current.columns: continue

            cdict = st.session_state.config[sheet].get(
                c_name, {"strategy": "keep", "params": {}}
            )
            current_strat = cdict.get("strategy", "keep")
            if current_strat not in STRATEGIES: current_strat = "keep"

            n_form = has_formula_in_column(formulas_current, c_name)
            label_prefix = "📐 " if n_form > 0 else "🔠 "
            label_suffix = f" — {STRATEGIES[current_strat]}"
            if n_form > 0:
                label_suffix += f" · {n_form} formule(s)"

            with st.expander(f"{label_prefix}{c_name}{label_suffix}", expanded=False):
                if n_form > 0:
                    st.info(f"Cette colonne contient {n_form} formule(s) Excel. "
                            f"En 'keep', elles seront restaurées dans l'export.")
                    samples_in_col = [(r, f) for (r, c), f in formulas_current.items()
                                      if c == c_name][:3]
                    for r, f in samples_in_col:
                        st.code(f"L{r+2}: {f}", language="text")

                new_strat = st.selectbox(
                    "Stratégie",
                    options=list(STRATEGIES.keys()),
                    format_func=lambda x: STRATEGIES[x],
                    index=list(STRATEGIES.keys()).index(current_strat),
                    key=f"strat_{sheet}_{c_name}",
                )

                try:
                    series = df_current[c_name]
                    samples = [str(v) for v in series.dropna().head(3) if not is_missing(v)]
                    if samples:
                        st.caption("Exemples : " + " · ".join(s[:20] for s in samples))
                except Exception:
                    st.caption("(aperçu indisponible)")

                new_params = {}
                for p_name, p_label, p_default, p_type in STRATEGY_PARAMS[new_strat]:
                    smart = smart_default(c_name, new_strat, all_cols, p_name)
                    fallback = smart if smart is not None else p_default
                    stored = cdict.get("params", {}).get(p_name)
                    if stored in (None, ""):
                        initial = str(fallback) if fallback != "" else ""
                    else:
                        initial = str(stored)
                    val = st.text_input(
                        p_label, value=initial,
                        key=f"param_{sheet}_{c_name}_{p_name}_{new_strat}",
                    )
                    new_params[p_name] = val

                st.session_state.config[sheet][c_name] = {
                    "strategy": new_strat, "params": new_params,
                }


# ─── COLONNE DROITE ───
with col_droite:
    st.subheader("👁️ Aperçu en direct")

    with st.expander("🔍 Filtres (vide = tout afficher)"):
        filter_cols = st.multiselect(
            "Colonnes à filtrer",
            list(df_current.columns),
            key=f"filter_cols_{sheet}",
        )
        df_filtered = df_current.copy()
        for fcol in filter_cols:
            if fcol not in df_filtered.columns: continue
            try:
                unique_vals = sorted({
                    str(v) for v in df_filtered[fcol].dropna() if not is_missing(v)
                })
            except Exception:
                unique_vals = []
            selected = st.multiselect(
                f"Valeurs pour '{fcol}'", unique_vals, default=[],
                key=f"filter_{sheet}_{fcol}",
            )
            if selected:
                df_filtered = df_filtered[
                    df_filtered[fcol].astype(str).isin(selected)
                ]

    if len(df_filtered) == 0:
        st.warning("Aucune ligne ne correspond aux filtres.")
        st.stop()

    max_rows = len(df_filtered)
    if max_rows > 1:
        row_range = st.select_slider(
            "Plage de lignes",
            options=list(range(max_rows + 1)),
            value=(0, min(100, max_rows)),
            key=f"row_range_{sheet}",
        )
        df_snippet = df_filtered.iloc[row_range[0]:row_range[1]].copy()
    else:
        df_snippet = df_filtered.copy()

    try:
        engine = PseudoEngine(copy.deepcopy(st.session_state.map_data))
        df_preview_full = engine.apply_sheet(
            df_snippet, sheet, st.session_state.config[sheet]
        )
    except Exception as e:
        st.error(f"Erreur d'aperçu : {e}")
        st.stop()

    # Pagination flèches + slider indépendants
    max_cols = len(df_preview_full.columns)
    cols_per_page = 5
    page_key = f"col_page_{sheet}"
    offset_key = f"col_offset_{sheet}"
    if page_key not in st.session_state: st.session_state[page_key] = 0
    if offset_key not in st.session_state: st.session_state[offset_key] = 0

    n_pages = max(1, (max_cols + cols_per_page - 1) // cols_per_page)
    if st.session_state[page_key] >= n_pages:
        st.session_state[page_key] = 0

    if max_cols > cols_per_page:
        nav1, nav2, nav3 = st.columns([1, 3, 1], vertical_alignment="center")
        with nav1:
            if st.button("⬅️ Page", disabled=st.session_state[page_key] <= 0,
                         use_container_width=True, key=f"prev_{sheet}"):
                st.session_state[page_key] -= 1
                st.rerun()
        with nav2:
            max_offset = max(0, min(cols_per_page - 1, max_cols - 1))
            if max_offset > 0:
                new_offset = st.slider(
                    "Décalage fin", 0, max_offset,
                    min(st.session_state[offset_key], max_offset),
                    label_visibility="collapsed",
                    key=f"offsetslider_{sheet}",
                )
                st.session_state[offset_key] = new_offset
            else:
                st.session_state[offset_key] = 0
            page = st.session_state[page_key]
            offset = st.session_state[offset_key]
            start = min(page * cols_per_page + offset, max(0, max_cols - 1))
            end = min(start + cols_per_page, max_cols)
            st.markdown(
                f"<div style='text-align:center;color:#888'>"
                f"Page <b>{page+1}/{n_pages}</b> · décalage <b>+{offset}</b> · "
                f"colonnes <b>{start+1}–{end}</b> sur {max_cols}"
                f"</div>", unsafe_allow_html=True,
            )
        with nav3:
            if st.button("Page ➡️",
                         disabled=st.session_state[page_key] >= n_pages - 1,
                         use_container_width=True, key=f"next_{sheet}"):
                st.session_state[page_key] += 1
                st.rerun()
    else:
        st.session_state[page_key] = 0
        st.session_state[offset_key] = 0

    page = st.session_state[page_key]
    offset = st.session_state[offset_key]
    cs = min(page * cols_per_page + offset, max(0, max_cols - 1))
    df_preview_view = df_preview_full.iloc[:, cs:cs + cols_per_page]
    df_snippet_view = df_snippet.iloc[:, cs:cs + cols_per_page]

    tab_apres, tab_avant, tab_form = st.tabs([
        "✨ Après pseudonymisation",
        "📋 Original",
        f"📐 Formules ({n_formulas_total})",
    ])
    with tab_apres:
        if engine._warnings:
            for w in engine._warnings[:5]: st.warning(w)
            if len(engine._warnings) > 5:
                st.caption(f"…et {len(engine._warnings)-5} autres.")
        st.dataframe(safe_dataframe_for_display(df_preview_view),
                     use_container_width=True, height=600, hide_index=True)
    with tab_avant:
        st.dataframe(safe_dataframe_for_display(df_snippet_view),
                     use_container_width=True, height=600, hide_index=True)
    with tab_form:
        if not formulas_current:
            st.info("Aucune formule détectée dans cette feuille.")
        else:
            st.caption(f"{n_formulas_total} formule(s). Préservées si stratégie 'keep'.")
            df_form = pd.DataFrame([
                {"Ligne Excel": r + 2, "Colonne": c, "Formule": f,
                 "Stratégie": st.session_state.config[sheet].get(c, {}).get("strategy", "keep")}
                for (r, c), f in formulas_current.items()
            ]).sort_values(["Ligne Excel", "Colonne"])
            st.dataframe(df_form, use_container_width=True, hide_index=True, height=600)

    st.caption(
        "ℹ️ Aperçu = pseudos calculés sur la plage visible. "
        "L'export utilise tout le classeur (mêmes valeurs → mêmes pseudos)."
    )
